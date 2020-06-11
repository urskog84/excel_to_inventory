import openpyxl
import yaml


wb = openpyxl.load_workbook('ip-plan.xlsx')

sheet = wb['IP Plan 2.0']

hedder = []

for i in range(1, sheet.max_column+1):
    title = sheet.cell(row=1, column=i).value
    hedder.append(title)

for i in range(2, sheet.max_row):
    obj = {}
    for j in range(1, sheet.max_column+1):
        k = hedder[j-1]
        v = sheet.cell(row=i, column=j).value
        d = {k : v}
        obj.update(d)
        filename = obj['Hostname']
        with open(f'inventory/host_vars/{filename}.yml', 'w') as file:
            documents = yaml.dump(obj, file)